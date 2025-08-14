from openpyxl import load_workbook as op
import time
import win32com.client
import pythoncom
import os
import variable 
import variable.config as config

def get_paths():
    home_dir = os.path.expanduser("~")
    base_dir = os.path.join(home_dir, "바탕 화면", "검사서")
    excel_dir = os.path.join(base_dir, "검사서 excel")
    pdf_dir = os.path.join(base_dir, "검사서 pdf")
    os.makedirs(excel_dir, exist_ok=True)
    os.makedirs(pdf_dir, exist_ok=True)
    
    now_str = time.strftime("%Y.%m.%d %H_%M_%S")
    file_name = f"{now_str} 검사서"
    
    excel_path = os.path.join(excel_dir, f"{file_name}.xlsx")
    pdf_path = os.path.join(pdf_dir, f"{file_name}.pdf")
    
    return excel_path, pdf_path

def fill_template(template_path, data_dict, cell_map, save_path):
    wb = op(template_path)
    ws = wb.active

    for key, values in data_dict.items():
        if key in cell_map:
            cells = cell_map[key]
            for i, value in enumerate(values):
                if i < len(cells):
                    ws[cells[i]] = value

    def fill_profile(ws, keys, values, profile_map):
        for key, value in zip(keys, values):
            if key in profile_map:
                ws[profile_map[key]] = value

    fill_profile(ws, config.labels, variable.profile_list, config.profile_map)

    
    wb.save(save_path)

def excel_to_pdf(excel_path, pdf_path):
    pythoncom.CoInitialize()
    excel = win32com.client.Dispatch("Excel.Application")
    excel.Visible = False
    wb = excel.Workbooks.Open(excel_path)
    wb.ExportAsFixedFormat(0, pdf_path)
    wb.Close()
    excel.Quit()
